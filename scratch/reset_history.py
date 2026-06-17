import openpyxl
from pathlib import Path

def reset_excel_history(file_path):
    print(f"Opening workbook: {file_path}")
    if not Path(file_path).exists():
        print("Error: File not found.")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        
        # 1. Clear Fallaways and Moveaways sheets
        for sheet_name in ["Fallaways", "Moveaways"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"Clearing sheet: {sheet_name}")
                # Delete all rows except the first (header)
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
        
        # 2. Clear sections in Combined sheet
        if "Combined" in wb.sheetnames:
            ws = wb["Combined"]
            print("Cleaning 'Combined' sheet sections...")
            
            for section_label in ["FALLAWAYS", "MOVEAWAYS"]:
                section_row = None
                # Find the label
                for r in range(1, 500):
                    for c in [1, 2, 3]:
                        val = str(ws.cell(row=r, column=c).value or "").strip().upper()
                        if section_label in val:
                            section_row = r
                            break
                    if section_row: break
                
                if section_row:
                    print(f"Found {section_label} at row {section_row}. Clearing entries below...")
                    # Clear up to 50 rows below the label, or until another major section/TOTAL
                    for r in range(section_row + 1, section_row + 51):
                        # Stop if we hit the other section or TOTAL
                        cell_val = str(ws.cell(row=r, column=2).value or "").strip().upper()
                        if any(x in cell_val for x in ["TOTAL", "FALLAWAYS", "MOVEAWAYS"]):
                            break
                        
                        # Clear columns A-G
                        for c in range(1, 8):
                            ws.cell(row=r, column=c, value=None)
                            ws.cell(row=r, column=c).border = None
                            ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(fill_type=None)

        wb.save(file_path)
        print("Successfully reset Excel history.")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    template = "C:/Users/User/Downloads/April 19, 2026 Contribution Report.xlsx"
    reset_excel_history(template)
