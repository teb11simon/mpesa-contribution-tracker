"""
Excel Generator
Creates Excel contribution reports by updating a master template workbook.
Supports formula preservation and automated sheet linking.
"""

import logging
import re
import traceback
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Set up logging
logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Generate Excel contribution reports using a template"""

    def __init__(self, template_path: Optional[str] = None):
        self.template_path = template_path
        self.workbook: Optional[Workbook] = None
        self.styles = self._create_styles()

    def _create_styles(self) -> Dict:
        """Create standard cell styles used in the ICC reports"""
        return {
            'header': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'title': Font(name='Calibri', size=14, bold=True),
            'bold': Font(name='Calibri', size=11, bold=True),
            'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'currency': '#,##0.00',
            'fill_gave': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
            'fill_under': PatternFill(start_color='E26B0A', end_color='E26B0A', fill_type='solid'),
            'fill_over': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
            'fill_missed': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'fill_na': PatternFill(start_color='002060', end_color='002060', fill_type='solid'),
            'font_white': Font(name='Calibri', size=11, color='FFFFFF')
        }

    @staticmethod
    def _get_report_sunday(report_date: datetime) -> datetime:
        """Returns report_date if Sunday, otherwise the most recent Sunday."""
        weekday = report_date.weekday()   # 0=Mon … 6=Sun
        if weekday == 6:
            return report_date.replace(hour=0, minute=0, second=0, microsecond=0)
        return (report_date - timedelta(days=(weekday + 1) % 7)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )

    def load_template(self, path: str):
        """Load the master/reference workbook and mark it as the source"""
        if not Path(path).exists():
            raise FileNotFoundError(f"Template file not found: {path}")
        self.template_path = path
        self.workbook = openpyxl.load_workbook(path)
        logger.info(f"Loaded source template from {path} (Original will be preserved)")

    def create_report(
        self,
        mpesa_transactions: List[Dict],
        cash_entries: List[Dict],
        report_date: datetime,
        output_path: str,
        template_path: Optional[str] = None
    ) -> str:
        """
        Main entry point to generate the weekly report.
        1. Loads template
        2. Extracts existing member list
        3. Creates new Sunday sheet (e.g. 'May 10')
        4. Updates Combined master sheet
        5. Updates Missions/Benevolence logs
        """
        if template_path:
            self.load_template(template_path)
        elif not self.workbook:
            # Fallback to creating a new workbook if no template provided
            self.workbook = Workbook()
            self._initialize_empty_workbook()

        # 1. Format Sunday Name (e.g., "May 3")
        sunday_name = report_date.strftime("%b %d").replace(" 0", " ")

        # 2. Extract Member List from Combined
        members = self._get_members_from_combined()

        # 3. Create/Update Sunday Sheet
        self._create_sunday_sheet(sunday_name, report_date, mpesa_transactions, cash_entries, members)

        # 4. Update Combined Sheet with new column
        self._update_combined_sheet(sunday_name, mpesa_transactions, cash_entries, members)

        # 5. Update Logs (Missions, Benevolence, Attendance)
        self._update_logs(report_date, mpesa_transactions, cash_entries)

        # 6. Save
        self.workbook.save(output_path)
        return output_path

    def _get_members_from_combined(self) -> List[Dict]:
        """
        Extracts member details from the 'Combined' sheet.
        Capture all columns to ensure alignment during sorting.
        """
        if "Combined" not in self.workbook.sheetnames:
            logger.warning("'Combined' sheet not found. Returning empty member list.")
            return []

        ws = self.workbook["Combined"]
        members = []
        max_col = ws.max_column
        
        # Start from row 2 (skipping header)
        for row in range(2, ws.max_row + 10):
            row_vals = [str(ws.cell(row=row, column=c).value or "").strip().lower() for c in range(1, 8)]
            
            # Stop ONLY if we see "TOTAL" anywhere in this row
            if any("total" in v for v in row_vals):
                break
            
            first_name = ws.cell(row=row, column=5).value
            last_name = ws.cell(row=row, column=6).value
            ministry = ws.cell(row=row, column=4).value
            
            if first_name or last_name or ministry:
                # Capture ALL cell data in the row for sorting preservation
                row_data = []
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=c)
                    row_data.append({
                        'value': cell.value,
                        'font': cell.font,
                        'border': cell.border,
                        'fill': cell.fill,
                        'number_format': cell.number_format,
                        'alignment': cell.alignment
                    })

                members.append({
                    'row_index': row,
                    'region': ws.cell(row=row, column=2).value,
                    'bible_talk': ws.cell(row=row, column=3).value,
                    'ministry': ministry,
                    'first_name': str(first_name).strip() if first_name else "",
                    'last_name': str(last_name).strip() if last_name else "",
                    'pledge': ws.cell(row=row, column=7).value or 0,
                    'is_bold': ws.cell(row=row, column=5).font.bold if first_name else False,
                    'full_row_data': row_data # Store the full row for sorting
                })
            
        return members

    def _save_to_template(self):
        """Saves the current state to a NEW file in the output folder to preserve the original"""
        if self.workbook:
            try:
                # Ensure output directory exists
                output_dir = Path("output")
                output_dir.mkdir(exist_ok=True)
                
                # Create a name based on the original template
                original_name = Path(self.template_path).stem if self.template_path else "Ledger"
                new_filename = f"{original_name}_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                save_path = output_dir / new_filename
                
                self.workbook.save(str(save_path))
                logger.info(f"Saved changes to NEW file: {save_path}")
                return str(save_path)
            except Exception as e:
                logger.error(f"Failed to save updated ledger: {e}")
                raise RuntimeError(f"Could not save changes. Error: {e}")
        return None

    def add_member(self, region, bible_talk, ministry, first_name, last_name, pledge):
        """Adds a new member to the Combined sheet before the TOTAL row"""
        if "Combined" not in self.workbook.sheetnames: return
        ws = self.workbook["Combined"]
        
        # Find the insertion point (the TOTAL row or the first empty name row)
        total_row = None
        for r in range(2, ws.max_row + 5):
            row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 8)]
            if any("total" in v for v in row_vals):
                total_row = r
                break
        
        if not total_row: total_row = ws.max_row + 1
        
        ws.insert_rows(total_row)
        
        # Calculate next number for Column A
        last_num = 0
        try:
            prev_val = ws.cell(row=total_row - 1, column=1).value
            if isinstance(prev_val, (int, float)): last_num = int(prev_val)
            elif str(prev_val).isdigit(): last_num = int(prev_val)
        except: pass
        
        # Fill data
        ws.cell(row=total_row, column=1, value=last_num + 1)
        ws.cell(row=total_row, column=2, value=region)
        ws.cell(row=total_row, column=3, value=bible_talk)
        ws.cell(row=total_row, column=4, value=ministry)
        ws.cell(row=total_row, column=5, value=first_name)
        ws.cell(row=total_row, column=6, value=last_name)
        
        try:
            p_val = float(re.sub(r'[^\d.]', '', str(pledge))) if pledge else 0
            ws.cell(row=total_row, column=7, value=p_val)
            ws.cell(row=total_row, column=7).number_format = self.styles['currency']
        except:
            ws.cell(row=total_row, column=7, value=pledge)
            
        for c in range(1, 8):
            ws.cell(row=total_row, column=c).border = self.styles['border']
            
        logger.info(f"Added member: {first_name} {last_name}")
        return self.sort_members() # This will return the new save path

    def remove_member(self, first_name, removal_type="fallaway"):
        """Removes member(s) matching the name and moves to Fallaways or Moveaways"""
        if "Combined" not in self.workbook.sheetnames: return
        ws = self.workbook["Combined"]
        
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            f_name = str(ws.cell(row=r, column=5).value or "").strip().lower()
            l_name = str(ws.cell(row=r, column=6).value or "").strip().lower()
            if f_name == first_name.lower() or l_name == first_name.lower():
                rows_to_delete.append(r)
        
        if not rows_to_delete:
            logger.warning(f"Member with name '{first_name}' not found.")
            return False
            
        # Capture data for archiving before deletion
        archive_data = []
        for r in rows_to_delete:
            archive_data.append({
                'date': datetime.now().strftime("%Y-%m-%d"),
                'region': ws.cell(row=r, column=2).value,
                'bible_talk': ws.cell(row=r, column=3).value,
                'ministry': ws.cell(row=r, column=4).value,
                'first_name': ws.cell(row=r, column=5).value,
                'last_name': ws.cell(row=r, column=6).value,
                'pledge': ws.cell(row=r, column=7).value
            })
            
        # Delete rows from bottom to top
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)
            
        # 1. Add to separate History Sheet (Fallaways or Moveaways)
        sheet_name = "Fallaways" if removal_type == "fallaway" else "Moveaways"
        if sheet_name not in self.workbook.sheetnames:
            fs = self.workbook.create_sheet(sheet_name)
            headers = ["Date Removed", "Region", "Bible Talk", "Ministry", "First Name", "Last Name", "Pledge"]
            for c, h in enumerate(headers, 1):
                cell = fs.cell(row=1, column=c, value=h)
                cell.font = self.styles['header']
                cell.border = self.styles['border']
        else:
            fs = self.workbook[sheet_name]
            
        for f in archive_data:
            next_row = fs.max_row + 1
            fs.cell(row=next_row, column=1, value=f['date'])
            fs.cell(row=next_row, column=2, value=f['region'])
            fs.cell(row=next_row, column=3, value=f['bible_talk'])
            fs.cell(row=next_row, column=4, value=f['ministry'])
            fs.cell(row=next_row, column=5, value=f['first_name'])
            fs.cell(row=next_row, column=6, value=f['last_name'])
            c7 = fs.cell(row=next_row, column=7, value=f['pledge'])
            c7.number_format = self.styles['currency']
            for c in range(1, 8):
                fs.cell(row=next_row, column=c).border = self.styles['border']
        
        # 2. Add to Combined sheet under the correct section (FALLAWAYS or MOVEAWAYS)
        section_label = "FALLAWAYS" if removal_type == "fallaway" else "MOVEAWAYS"
        section_start_row = None
        for r in range(2, 400):
            for c in [1, 2, 3]:
                val = str(ws.cell(row=r, column=c).value or "").strip().upper()
                if section_label in val:
                    section_start_row = r
                    break
            if section_start_row: break
        
        if section_start_row:
            # Find first truly empty row (Start directly after the header)
            target_row = None
            for r in range(section_start_row + 1, section_start_row + 50):
                # Check column E (First Name) to see if row is occupied
                if not ws.cell(row=r, column=5).value:
                    target_row = r
                    break
            
            if target_row:
                today = datetime.now().strftime("%Y-%m-%d")
                for f in archive_data:
                    ws.cell(row=target_row, column=1, value=today) # A: Date
                    ws.cell(row=target_row, column=2, value=f['region']) # B: Region
                    ws.cell(row=target_row, column=3, value=f['bible_talk']) # C: BT
                    ws.cell(row=target_row, column=4, value=f['ministry']) # D: Ministry
                    ws.cell(row=target_row, column=5, value=f['first_name']) # E: FN
                    ws.cell(row=target_row, column=6, value=f['last_name']) # F: LN
                    c7 = ws.cell(row=target_row, column=7, value=f['pledge']) # G: Pledge
                    c7.number_format = self.styles['currency']
                    c7.font = self.styles['bold']
                    
                    # Apply borders to all columns A-G
                    for c in range(1, 8):
                        ws.cell(row=target_row, column=c).border = self.styles['border']
                        if c == 1: ws.cell(row=target_row, column=c).alignment = self.styles['center']
                    target_row += 1
            
        logger.info(f"Member(s) moved to {section_label}.")
        return self.sort_members() # Returns new save path

    def sort_members(self):
        """Sorts members in Combined sheet alphabetically by Last Name (F) and re-numbers A, preserving all columns"""
        if "Combined" not in self.workbook.sheetnames: return
        ws = self.workbook["Combined"]
        
        # 1. Extract all current members with full row data
        members = self._get_members_from_combined()
        if not members: return
        
        # Filter: Only keep rows that have at least one NAME (ignore ministry-only rows)
        members = [m for m in members if (m['first_name'].strip() or m['last_name'].strip())]
        
        # 2. Sort by last name (if last name is empty, use first name)
        members.sort(key=lambda x: (x['last_name'].lower() or x['first_name'].lower() or "zzz"))
        
        # 3. Find the TOTAL row to define the clearable area
        actual_total_row = ws.max_row + 1
        for r in range(2, ws.max_row + 10):
            row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 8)]
            if any("total" in v for v in row_vals):
                actual_total_row = r
                break
        
        max_col = ws.max_column
        
        # 4. Clear the member area (All columns)
        for r in range(2, actual_total_row):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment()
        
        # 5. Write back the cleaned, sorted list with full row data
        from copy import copy
        for i, m in enumerate(members):
            row = i + 2
            full_data = m.get('full_row_data', [])
            
            for c_idx, data in enumerate(full_data, 1):
                cell = ws.cell(row=row, column=c_idx)
                
                # Column A: Always update to new sequential number
                if c_idx == 1:
                    cell.value = i + 1
                    cell.alignment = self.styles['center']
                else:
                    cell.value = data['value']
                    if data['alignment']: cell.alignment = copy(data['alignment'])
                
                # Restore styles
                if data['font']: cell.font = copy(data['font'])
                if data['border']: cell.border = copy(data['border'])
                if data['fill']: cell.fill = copy(data['fill'])
                if data['number_format']: cell.number_format = data['number_format']
            
            # Ensure Column A has border if full_data was shorter than expected
            ws.cell(row=row, column=1).border = self.styles['border']
                
        logger.info(f"Sorted and preserved {len(members)} members with all column data.")
        
        # --- SAFE SYNC: Only fix the titles if they are broken ---
        for section in ["FALLAWAYS", "MOVEAWAYS"]:
            for r in range(20, 100):
                for c in [1, 2, 3]:
                    val = str(ws.cell(row=r, column=c).value or "").strip().upper()
                    if val == "LA" or (section in val and val != section):
                        ws.cell(row=r, column=c, value=section)
                        break
        
        return self._save_to_template()

    def _create_sunday_sheet(self, name: str, date: datetime, mpesa: List[Dict], cash: List[Dict], members: List[Dict]):
        """Creates the detailed Sunday contribution sheet"""
        if name in self.workbook.sheetnames:
            # If sheet exists, remove it or rename? For now, we overwrite.
            del self.workbook[name]
            
        ws = self.workbook.create_sheet(name, 0) # Insert at beginning
        
        # Header Block
        ws['A3'] = "Nairobi ICC"
        ws['A3'].font = self.styles['title']
        
        ws['A5'] = "Date:"
        ws['B5'] = date.strftime("%Y-%m-%d")
        ws['C5'] = "Contribution"
        ws['D5'] = "=SUM(F:F)+SUM(K:K)"
        
        ws['A6'] = "Date:"
        ws['B6'] = date.strftime("%Y-%m-%d")
        ws['C6'] = "Missions"
        ws['D6'] = "=SUM(G:G)+SUM(N:N)"
        
        # Headers Row
        headers = ["#", "Ministry", "First Name", "Last Name", "Pledge in Ksh", "Contribution", "Missions", "Total"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=9, column=i, value=h)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']

        # Populate Members
        row_idx = 10
        for i, m in enumerate(members, 1):
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=m['ministry'])
            ws.cell(row=row_idx, column=3, value=m['first_name'])
            ws.cell(row=row_idx, column=4, value=m['last_name'])
            ws.cell(row=row_idx, column=5, value=m['pledge'])
            
            # These will be populated by the matching logic later
            # For now, we put 0 or placeholders
            ws.cell(row=row_idx, column=6, value=0) # Contribution
            ws.cell(row=row_idx, column=7, value=0) # Missions
            ws.cell(row=row_idx, column=8, value=f"=F{row_idx}+G{row_idx}") # Total
            
            row_idx += 1
            
        # Add Visitor Section
        ws.cell(row=row_idx + 1, column=10, value="Visitors' Contribution")
        ws.cell(row=row_idx + 1, column=10).font = self.styles['bold']

    def _update_summary_sheet(self, date_str: str, sunday_date: datetime):
        """Updates the Summary dashboard with new date and links to the latest Sunday sheet"""
        if "Summary" not in self.workbook.sheetnames: return
        ws = self.workbook["Summary"]
        
        # 0. Update Year in Title (e.g. "2025 Weekly Contribution & Benevolence")
        current_year = str(datetime.now().year)
        for r in range(1, 10): # Look in the first 10 rows
            for c in range(1, 10):
                val = str(ws.cell(row=r, column=c).value or "")
                if "Weekly Contribution & Benevolence" in val:
                    # Replace 4-digit year with current year
                    new_val = re.sub(r"\d{4}", current_year, val)
                    ws.cell(row=r, column=c, value=new_val)
                    break
        formatted_date = sunday_date.strftime("%#d %B %Y") # e.g. 6 May 2026
        e11_val = str(ws['E11'].value or "Given")
        if "Given" in e11_val:
            ws['E11'] = f"Given {formatted_date}"
        else:
            ws['E11'] = f"Given {formatted_date}"
            
        # 1b. Clear C18 (per user request)
        ws['C18'] = None
        ws['C18'].border = None

        # 1c. Update C8: Today if Sunday, else last Sunday
        now = datetime.now()
        # weekday() returns 6 for Sunday (0=Mon, 6=Sun)
        days_to_subtract = now.weekday() + 1 if now.weekday() != 6 else 0
        target_sunday = now if now.weekday() == 6 else (now - timedelta(days=days_to_subtract))
        ws['C8'] = target_sunday.strftime("%#d %B %Y")
        ws['C8'].alignment = self.styles['center']
        ws['C8'].font = self.styles['bold']
            
        # 2. Update F12 (Missions) -> SundayTab!D6
        ws['F12'] = f"='{date_str}'!D6"
        
        # 3. Update C12 (Contribution) -> SundayTab!D4
        ws['C12'] = f"='{date_str}'!D4"
        
        # 3a. Update D12 (Weekly Reference) -> SundayTab!D7
        ws['D12'] = f"='{date_str}'!D7"
        
        # 3b. Update G12, I12, and I13 to point to 'Missions' tab and local cells
        ws['G12'] = "='Missions'!C2"
        ws['I12'] = "='Missions'!C1"
        ws['I13'] = "=I12"
        
        # 4. Update ALL other formulas that reference a sheet tab
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    # EXCLUDE G12 and I12 from the global weekly update
                    if cell.coordinate in ['G12', 'I12']:
                        # Ensure they point to Missions if they don't already
                        if "'Missions'!" not in cell.value and "Missions!" not in cell.value:
                            # If they are currently #REF! or pointing elsewhere, fix them
                            formula = cell.value
                            new_formula = re.sub(r"('?[^'!]+'?)!", "'Missions'!", formula)
                            cell.value = new_formula
                        continue
                        
                    formula = cell.value
                    new_formula = re.sub(r"('?[^'!]+'?)!", f"'{date_str}'!", formula)
                    if new_formula != formula:
                        cell.value = new_formula

        logger.info(f"Summary dashboard healed and synchronized with '{date_str}'.")

    def _update_combined_sheet(self, date_str: str, mpesa_transactions: List[Dict], cash_entries: List[Dict], members: List[Dict]):
        """Adds a new column to Combined sheet and populates it"""
        if "Combined" not in self.workbook.sheetnames: return
        ws = self.workbook["Combined"]

        from openpyxl.styles import Border, Side, Font
        _thin        = Side(border_style='thin')
        _full_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        # ── Repair pass: backfill borders + red fill on ALL existing weekly ──
        # columns for any member row that has blank/borderless cells.
        # This fixes newer members who joined after historical columns were created.
        # Find the last weekly column by scanning row 1 for date headers (col 9+)
        last_weekly_col = 8
        for c in range(9, ws.max_column + 1):
            if ws.cell(row=1, column=c).value is not None:
                last_weekly_col = c
            else:
                break

        # Get pledge lookup for fill colouring during repair
        pledge_by_row = {}
        for r in range(2, ws.max_row + 1):
            first = ws.cell(row=r, column=5).value
            if first is None: break
            pledge_by_row[r] = ws.cell(row=r, column=7).value

        for r in range(2, ws.max_row + 1):
            first = ws.cell(row=r, column=5).value
            if first is None: break
            pledge = pledge_by_row.get(r)
            pledge_str = str(pledge or '').strip().lower()
            is_na = pledge_str in ('n/a', 'na', '')

            # ── Repair border on col 8 (Given) — no fill change ─────
            col8 = ws.cell(row=r, column=8)
            b8 = col8.border
            has8 = any(getattr(b8, s).border_style for s in ['left','right','top','bottom'] if getattr(b8,s))
            if not has8:
                col8.border = _full_border

            # ── Repair borders + fill on historical weekly cols (9+) ──
            for c in range(9, last_weekly_col + 1):
                cell = ws.cell(row=r, column=c)
                b    = cell.border
                has_border = any(
                    getattr(b, s).border_style
                    for s in ['left', 'right', 'top', 'bottom']
                    if getattr(b, s)
                )
                if not has_border:
                    cell.border = _full_border
                    if cell.value is None:
                        cell.value = 0
                    if is_na:
                        cell.fill = self.styles['fill_na']
                        cell.font = self.styles['font_white']
                    elif (cell.value or 0) == 0:
                        cell.fill = self.styles['fill_missed']
                        cell.font = self.styles['font_white']

        # Insert new column at I (column index 9)
        ws.insert_cols(9)

        # Set header
        header_cell = ws.cell(row=1, column=9, value=date_str)
        header_cell.font      = self.styles['bold']
        header_cell.alignment = self.styles['center']
        header_cell.border    = _full_border

        # ── Build a name → current row lookup by scanning Combined NOW ──
        name_to_row = {}
        for r in range(2, ws.max_row + 1):
            first = ws.cell(row=r, column=5).value
            last  = ws.cell(row=r, column=6).value
            if first is None:
                break
            key = (str(first).strip().lower(), str(last or '').strip().lower())
            name_to_row[key] = r

        # ── Write ALL member rows — zero or not — with full border + fill ──
        col_total = 0
        last_member_row = 1

        for m in members:
            first_key   = str(m.get('first_name', '')).strip().lower()
            last_key    = str(m.get('last_name',  '')).strip().lower()
            current_row = name_to_row.get((first_key, last_key))

            if current_row is None:
                logger.warning(f"Combined: could not find row for {m.get('first_name')} {m.get('last_name')} — skipping.")
                continue

            # Retrieve total give (contribution + missions) by reading absolute value directly from column H of weekly tab
            master_row = m.get('row_index', current_row)
            weekly_row = master_row + 8
            
            amount = 0
            try:
                ws_weekly = self.workbook[date_str]
                weekly_val = ws_weekly.cell(row=weekly_row, column=8).value
                # If it's an absolute value (float/int), use it. If it's a formula, we fallback to recalculating
                if isinstance(weekly_val, (int, float)):
                    amount = weekly_val
                else:
                    # Fallback calculation if the weekly sheet hasn't been saved as data-only yet
                    cont_total = sum(t.get('amount', 0) for t in mpesa_transactions if t.get('member_row') == master_row and str(t.get('category', '')).lower().startswith('cont'))
                    cont_total += sum(c.get('amount', 0) for c in cash_entries if c.get('member_row') == master_row and str(c.get('category', '')).lower().startswith('cont'))
                    miss_total = sum(t.get('amount', 0) for t in mpesa_transactions if t.get('member_row') == master_row and str(t.get('category', '')).lower().startswith('miss'))
                    miss_total += sum(c.get('amount', 0) for c in cash_entries if c.get('member_row') == master_row and str(c.get('category', '')).lower().startswith('miss'))
                    amount = cont_total + miss_total
            except Exception:
                amount = 0
            data_cell = ws.cell(row=current_row, column=9, value=amount)
            data_cell.border        = _full_border
            data_cell.number_format = self.styles['currency']

            # Conditional colour fill
            pledge_val = m.get('pledge', 0)
            try:
                if isinstance(pledge_val, str) and 'n/a' in pledge_val.lower():
                    data_cell.fill = self.styles['fill_na']
                    data_cell.font = self.styles['font_white']
                else:
                    p_num = float(re.sub(r'[^\d.]', '', str(pledge_val or 0)))
                    if amount == 0:
                        data_cell.fill = self.styles['fill_missed']
                        data_cell.font = self.styles['font_white']
                    elif amount == p_num:
                        data_cell.fill = self.styles['fill_gave']
                        data_cell.font = Font()
                    elif amount > p_num:
                        data_cell.fill = self.styles['fill_over']
                        data_cell.font = Font()
                    else:
                        data_cell.fill = self.styles['fill_under']
                        data_cell.font = self.styles['font_white']
            except:
                pass

            col_total += amount
            if current_row > last_member_row:
                last_member_row = current_row

        # ── Write TOTAL row ────────────────────────────────────────────
        for r in range(last_member_row + 1, last_member_row + 5):
            if "TOTAL" in str(ws.cell(row=r, column=2).value or "").upper():
                col_letter = get_column_letter(9)
                total_cell              = ws.cell(row=r, column=9, value=f"=SUM({col_letter}2:{col_letter}{last_member_row})")
                total_cell.font         = self.styles['bold']
                total_cell.number_format = self.styles['currency']
                total_cell.border       = _full_border
                break

    def generate_excel(self, date: datetime, mpesa: List[Dict], cash: List[Dict], output_path: str):
        """Generates the Excel report and updates all tracker sheets"""
        date_str = date.strftime("%b %d").replace(" 0", " ")
        members = self._get_members_from_combined()
        
        # 1. Create/Update Sunday Sheet
        # (This is handled via create_report_with_matches in the current workflow)
        pass

    def _update_attendance_sheet(self, sunday_date: datetime, men: int, women: int, children: int):
        """Adds a new attendance record to the Attendance sheet using the report date"""
        # Case-insensitive and space-trimmed sheet search
        attendance_sheet_name = next((s for s in self.workbook.sheetnames if s.strip().lower() == "attendance"), None)
        if not attendance_sheet_name:
            logger.warning("Attendance sheet not found in workbook.")
            return
            
        ws = self.workbook[attendance_sheet_name]
        
        # 1. Use the actual report date provided
        date_val = sunday_date.strftime("%#d-%b-%y") # e.g. 3-May-26
        
        # 2. Find first empty row (looking at Column A)
        target_row = 4
        for r in range(4, 1000):
            val = ws.cell(row=r, column=1).value
            if val is None or str(val).strip() == "":
                target_row = r
                break
        
        # 3. Write Data (Use actual date object for Excel consistency)
        ws.cell(row=target_row, column=1, value=sunday_date)
        ws.cell(row=target_row, column=2, value=men)
        ws.cell(row=target_row, column=3, value=women)
        ws.cell(row=target_row, column=4, value=children)
        ws.cell(row=target_row, column=6, value=f"=SUM(B{target_row}:D{target_row})") # Total in F
        
        # 4. Apply Formatting
        for c in [1, 2, 3, 4, 6]:
            cell = ws.cell(row=target_row, column=c)
            cell.border = self.styles['border']
            
            # Apply specific alignment and font per column
            if c == 1: # Date
                cell.number_format = 'd-mmm-yy'
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif c == 6: # Total
                cell.font = self.styles['bold']
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else: # Counts (Men, Women, Children)
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='right', vertical='center')

        logger.info(f"Attendance recorded for {date_val} in row {target_row} (Men: {men}, Women: {women}, Children: {children}).")
        print(f"DEBUG: Attendance recorded for {date_val} in row {target_row}")

    def _update_bible_talk_report(self, members: List[Dict], all_transactions: Optional[List[Dict]] = None):
        """Groups members by Bible Talk in a multi-column grid layout matching the dashboard style"""
        try:
            sheet_name = next((s for s in self.workbook.sheetnames if "Bible Talk" in s), None)
            if not sheet_name:
                logger.warning("Bible Talk report sheet not found.")
                return
            ws = self.workbook[sheet_name]
            logger.info(f"Updating Bible Talk report on sheet: {sheet_name}")
            
            # 1. Clear existing content (Broad clearing for the grid area)
            for row in range(1, 300):
                for col in range(1, 26): # Columns A to Z
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None: cell.value = None
                    if cell.fill.fill_type is not None: cell.fill = PatternFill(fill_type=None)
                    # Clear borders safely
                    cell.border = Border()
                    cell.font = Font(name='Calibri', size=11)
            
            # Reset merges properly to avoid MergedCell conflicts
            for merged_range in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged_range))

            # 2. Build Contribution-only totals per member
            contribution_by_row: Dict[int, float] = {}
            if all_transactions:
                for t in all_transactions:
                    category = (t.get('category') or '').strip().lower()
                    if category != 'contribution':
                        continue
                    member_row = t.get('member_row')
                    if member_row is None:
                        continue
                    contribution_by_row[member_row] = contribution_by_row.get(member_row, 0.0) + (t.get('amount', 0) or 0)

            # 3. Group members by Bible Talk
            bt_groups = {}
            for m in members:
                bt_raw = (m.get('bible_talk') or "Other").strip()
                # Title case for consistent headers (e.g., "Mashujaa")
                bt = bt_raw.title() if bt_raw.lower() != "other" else "Other"
                if not bt or bt.lower() == "none": bt = "Other"
                if bt not in bt_groups: bt_groups[bt] = []
                bt_groups[bt].append(m)
            
            logger.info(f"Grouped members into {len(bt_groups)} Bible Talk(s): {list(bt_groups.keys())}")
            print(f"DEBUG: Found {len(bt_groups)} Bible Talk groups: {list(bt_groups.keys())}")
                
            # 3. Grid Settings (3 blocks per row)
            col_anchors = [1, 8, 15] # A, H, O
            curr_bt_idx = 0
            
            # 4. Styles for the BT Dashboard
            bt_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            bt_header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            
            # 5. Write Groups
            for bt_name in sorted(bt_groups.keys()):
                logger.info(f"Writing block for Bible Talk: {bt_name} ({len(bt_groups[bt_name])} members)")
                col_start = col_anchors[curr_bt_idx % 3]
                # Calculate row offset based on previous blocks in the same column
                row_start = 1 + (curr_bt_idx // 3) * 35 
                
                # A. Main BT Header (Merged A-E, H-L, O-S approx)
                try:
                    header_cell = ws.cell(row=row_start, column=col_start, value=f"{bt_name} BT")
                    header_cell.font = bt_header_font
                    header_cell.fill = bt_header_fill
                    header_cell.alignment = self.styles['center']
                    ws.merge_cells(start_row=row_start, start_column=col_start, end_row=row_start, end_column=col_start+3)
                except Exception as merge_err:
                    logger.warning(f"Merge conflict for {bt_name}: {merge_err}")
                
                # B. Sub-headers
                sub_headers = ["Name", "Last Name", "Pledge", "Contribution"]
                for i, h in enumerate(sub_headers):
                    c = ws.cell(row=row_start+1, column=col_start+i, value=h)
                    c.font = self.styles['bold']
                    c.border = self.styles['border']
                    c.alignment = self.styles['center']
                
                # C. Fill Members
                sorted_m = sorted(bt_groups[bt_name], key=lambda x: (x['last_name'].lower() or x['first_name'].lower() or "zzz"))
                
                row_idx = row_start + 2
                for m in sorted_m:
                    # 1. Name
                    name_cell = ws.cell(row=row_idx, column=col_start, value=m['first_name'])
                    name_cell.border = self.styles['border']
                    if m.get('is_bold'): name_cell.font = self.styles['bold']
                    
                    # 2. Last Name
                    ln_cell = ws.cell(row=row_idx, column=col_start+1, value=m['last_name'])
                    ln_cell.border = self.styles['border']
                    if m.get('is_bold'): ln_cell.font = self.styles['bold']
                    
                    # 3. Pledge
                    p_cell = ws.cell(row=row_idx, column=col_start+2, value=m['pledge'])
                    p_cell.border = self.styles['border']
                    p_cell.number_format = self.styles['currency']
                    if m.get('is_bold'): p_cell.font = self.styles['bold']
                    
                    # 4. Contribution Box (Amount + Color)
                    amount = contribution_by_row.get(m.get('row_index'), 0.0)
                    color_cell = ws.cell(row=row_idx, column=col_start+3, value=amount if amount > 0 else "")
                    color_cell.border = self.styles['border']
                    color_cell.alignment = self.styles['center']
                    
                    # Apply Giving Color and Font
                    pledge_val = m.get('pledge', 0)
                    try:
                        if isinstance(pledge_val, str) and 'n/a' in pledge_val.lower():
                            color_cell.fill = self.styles['fill_na']
                            color_cell.font = self.styles['font_white']
                        else:
                            # Safer numeric conversion
                            clean_pledge = re.sub(r'[^\d.]', '', str(pledge_val or 0))
                            p_num = float(clean_pledge) if clean_pledge else 0.0
                            
                            if amount == 0:
                                color_cell.fill = self.styles['fill_missed']
                                color_cell.font = self.styles['font_white']
                            elif amount >= p_num: # Simplified check to match master
                                color_cell.fill = self.styles['fill_gave'] if amount == p_num else self.styles['fill_over']
                                color_cell.font = Font(name='Calibri', size=11, bold=m.get('is_bold', False))
                            else:
                                color_cell.fill = self.styles['fill_under']
                                color_cell.font = self.styles['font_white']
                    except Exception as coloring_err:
                        logger.warning(f"Coloring error for member {m.get('first_name')}: {coloring_err}")
                    
                    row_idx += 1
                
                # D. Add Legend at the bottom of the block
                legend_row = row_idx + 1
                legends = [
                    ("Gave", self.styles['fill_gave']),
                    ("Underpledged", self.styles['fill_under']),
                    ("Overpledged", self.styles['fill_over']),
                    ("Missed", self.styles['fill_missed'])
                ]
                for label, fill in legends:
                    ws.cell(row=legend_row, column=col_start, value=label).border = self.styles['border']
                    ws.cell(row=legend_row, column=col_start+3).fill = fill
                    ws.cell(row=legend_row, column=col_start+3).border = self.styles['border']
                    legend_row += 1
                    
                curr_bt_idx += 1
            logger.info("Bible Talk report updated successfully.")
            
        except Exception as e:
            error_trace = traceback.format_exc()
            logger.error(f"Error updating Bible Talk report: {e}\n{error_trace}")
            print(f"DEBUG ERROR: Bible Talk Report failed -> {e}")
            # Show a temporary error in console to help debug
            print(f"TRACEBACK: {error_trace}")

    def finalize_report(self, date: datetime, mpesa: List[Dict], cash: List[Dict], members: List[Dict], output_path: str, attendance: Dict = None, all_transactions: Optional[List[Dict]] = None):
        """Called after the main report is generated to sync the other sheets"""
        date_str = date.strftime("%b %d").replace(" 0", " ")
        
        # Update Combined Sheet
        self._update_combined_sheet(date_str, mpesa, cash, members)
        
        # Update Summary Dashboard
        self._update_summary_sheet(date_str, date)
        
        # Update Attendance if provided
        if attendance:
            self._update_attendance_sheet(
                date,
                attendance.get('men', 0),
                attendance.get('women', 0),
                attendance.get('children', 0)
            )
            
        # Update Bible Talk Report
        self._update_bible_talk_report(members, all_transactions=all_transactions)

        # Update Income & Exp tab with all expense transactions
        if all_transactions:
            self._update_income_expense_tab(all_transactions)

        # Append a single "Contribution Collected" income row summing all income
        if all_transactions:
            total_contribution = sum(
                abs(t.get('amount', 0) or 0)
                for t in all_transactions
                if (t.get('category') or '').strip().lower() == 'contribution'
            )
            if total_contribution > 0:
                self._write_contribution_income_row(date, total_contribution)

        # Update Missing Contribution tab
        self._update_missing_contro_tab(date, members, all_transactions=all_transactions)

        # Update Missions tab
        if all_transactions:
            self._update_missions_tab(all_transactions)

        # Update Benevolence tab
        if all_transactions:
            self._update_benevolence_tab(all_transactions, date)

        # Create/update Breakdown of Available Funds tab
        self._create_breakdown_tab(date, all_transactions=all_transactions, members=members)
        
        # Save finally
        self.workbook.save(output_path)
        logger.info(f"Final sync completed for {date_str}")

    def _update_income_expense_tab(self, all_transactions: List[Dict]):
        """
        Appends EXPENSE-ONLY transactions to the 'Income & Exp' sheet.

        Only Paid Out transactions are written here. Contribution and Missions
        income are intentionally excluded — they are tracked in their own sheets.

        Sheet structure (row 5 = headers, data from row 6):
          A = S/No.  (formula =COUNTBLANK($S$5:Sn))
          B = Description
          C = Income / Expense  ← dropdown: "Income,Expense"
          D = Date
          E = Category          ← dropdown: expense category list
          F = Amount (KES)      — always negative for expenses
          G = Running Balance   (formula =G(prev)+F(curr))
          H = Reference Number  (M-Pesa receipt no. — used for deduplication)
        """
        if "Income & Exp" not in self.workbook.sheetnames:
            logger.warning("'Income & Exp' sheet not found — skipping.")
            return

        ws = self.workbook["Income & Exp"]

        # ── 1. Filter: keep only expense-categorised transactions ─────────
        # We check the user-assigned category, NOT the M-Pesa type.
        # A "Transaction Charge" that arrived as a Paid In is still an expense.
        # Anything the user categorised as Contribution, Missions, or Benevolence
        # is tracked in its own sheet and excluded here.
        INCOME_CATS = {'contribution', 'missions', 'benevolence'}

        expense_txns = []
        for t in all_transactions:
            category = (t.get('category') or '').strip().lower()
            if category in INCOME_CATS or not category:
                continue
            # Also skip the "Balance brought forward" sentinel row if present
            if 'balance' in category:
                continue
            expense_txns.append(t)

        if not expense_txns:
            logger.info("Income & Exp: no new expense transactions to append.")
            return

        # ── 2. Find the last real data row ───────────────────────────────
        last_row = 6
        for r in range(6, ws.max_row + 1):
            b = ws.cell(row=r, column=2).value
            if b is not None and str(b).strip() not in ("", "."):
                last_row = r

        # ── 3. Collect existing dedup keys ────────────────────────────────
        # M-Pesa reuses the same receipt number for a transaction AND its
        # associated charge (e.g. withdrawal + withdrawal charge share one ID).
        # Use receipt_no + amount as the composite key so both rows survive.
        existing_keys = set()
        for r in range(6, last_row + 1):
            ref = ws.cell(row=r, column=8).value        # col H = receipt no.
            amt = ws.cell(row=r, column=6).value        # col F = amount
            if ref:
                existing_keys.add(f"{str(ref).strip()}|{amt}")

        # ── 4. Inspect existing dropdowns so we can extend them if needed ───
        # C6:C566  — "Income,Expense" dropdown
        # E7:E267 and E268:E566 — category dropdown (split in two ranges)
        # New rows land beyond current last data row (≤173) and well within the
        # existing ceilings (566), so extension is rarely needed — handled
        # defensively in case the sheet grows large over the year.
        dv_inc_exp  = None    # col C dropdown
        dv_cat_low  = None    # col E dropdown with the lower start row
        dv_cat_high = None    # col E dropdown with the higher start row

        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list':
                continue
            formula   = dv.formula1 or ''
            sqref_str = str(dv.sqref)
            if 'Income' in formula and sqref_str.startswith('C'):
                dv_inc_exp = dv
            elif sqref_str.startswith('E'):
                try:
                    start_row = int(''.join(c for c in sqref_str.split(':')[0] if c.isdigit()))
                except (ValueError, IndexError):
                    continue
                if dv_cat_low is None:
                    dv_cat_low = (start_row, dv)
                elif start_row < dv_cat_low[0]:
                    dv_cat_high = dv_cat_low
                    dv_cat_low  = (start_row, dv)
                else:
                    dv_cat_high = (start_row, dv)

        # Ensure the two new categories are present in both col E dropdowns
        NEW_CATS = ["Benevolence Expense", "Benevolence Expense Transaction Charge"]
        for dv_tuple in [dv_cat_low, dv_cat_high]:
            if dv_tuple is None:
                continue
            dv_obj = dv_tuple[1]
            existing = dv_obj.formula1 or ''
            # Strip surrounding quotes if present
            clean = existing.strip('"')
            items = [i.strip() for i in clean.split(',')]
            added = False
            for cat in NEW_CATS:
                if cat not in items:
                    items.append(cat)
                    added = True
            if added:
                dv_obj.formula1 = '"' + ','.join(items) + '"'

        def _extend_dv(dv_obj, new_last_row: int):
            """Widen a DataValidation sqref ceiling to cover new_last_row."""
            if dv_obj is None:
                return
            sqref_str = str(dv_obj.sqref)
            parts = sqref_str.split(':')
            if len(parts) != 2:
                return
            col_letter = ''.join(c for c in parts[1] if c.isalpha())
            try:
                current_end = int(''.join(c for c in parts[1] if c.isdigit()))
            except ValueError:
                return
            if new_last_row > current_end:
                dv_obj.sqref = f"{parts[0]}:{col_letter}{new_last_row}"

        # ── 5. Copy style helper ──────────────────────────────────────────
        def _copy_style(src_row: int, dst_row: int, col: int):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=dst_row, column=col)
            if src.has_style:
                dst.font      = src.font.copy()
                dst.border    = src.border.copy()
                dst.alignment = src.alignment.copy()

        # ── 6. Sort by date ───────────────────────────────────────────────
        def _tx_date(t):
            d = t.get('date')
            return d if isinstance(d, datetime) else datetime.min

        sorted_txns = sorted(expense_txns, key=_tx_date)

        # ── 7. Append each new expense row ────────────────────────────────
        new_count = 0
        for t in sorted_txns:
            ref = str(t.get('receipt_no') or t.get('reference_no') or '').strip()
            raw_amount = t.get('amount', 0) or 0
            
            # Determine if it's Income or Expense based on transaction type
            tx_type = t.get('type') or t.get('transaction_type') or "Paid Out"
            if tx_type == "Paid In":
                amount = abs(raw_amount)
                inc_exp_val = "Income"
            else:
                amount = -abs(raw_amount)
                inc_exp_val = "Expense"

            # Composite dedup key: receipt_no + amount
            dedup_key = f"{ref}|{amount}" if ref else ''
            if dedup_key and dedup_key in existing_keys:
                continue

            description = (
                t.get('details') or
                t.get('description') or
                t.get('sender_name') or
                t.get('name') or
                ''
            )
            category = t.get('category') or inc_exp_val
            tx_date  = t.get('date') or datetime.now()

            new_row = last_row + 1

            # Col A: auto-number formula matching existing pattern
            s_ref_row = new_row - 2
            ws.cell(row=new_row, column=1).value = f"=COUNTBLANK($S$5:S{s_ref_row})"

            ws.cell(row=new_row, column=2).value = description          # B
            ws.cell(row=new_row, column=3).value = inc_exp_val          # C

            date_cell = ws.cell(row=new_row, column=4)                  # D
            date_cell.value         = tx_date
            date_cell.number_format = 'DD-MMM-YY'

            ws.cell(row=new_row, column=5).value = category             # E

            amt_cell = ws.cell(row=new_row, column=6)                   # F
            amt_cell.value         = amount
            amt_cell.number_format = '#,##0.00'

            ws.cell(row=new_row, column=7).value = f"=G{last_row}+F{new_row}"  # G

            ws.cell(row=new_row, column=8).value = ref if ref else None # H

            # Mirror borders/fonts from the previous row
            for col in range(1, 9):
                _copy_style(last_row, new_row, col)

            if dedup_key:
                existing_keys.add(dedup_key)
            last_row  = new_row
            new_count += 1

        # ── 8. Extend dropdowns to cover newly added rows ─────────────────
        if new_count > 0:
            if dv_inc_exp:
                _extend_dv(dv_inc_exp, last_row)
            if dv_cat_high:
                _extend_dv(dv_cat_high[1], last_row)
            if dv_cat_low:
                _extend_dv(dv_cat_low[1], last_row)

        logger.info(f"Income & Exp: appended {new_count} new expense transaction(s).")

    def _write_contribution_income_row(self, report_date: datetime, total_amount: float):
        """
        Appends a single 'CONTRIBUTION COLLECTED' income row to 'Income & Exp'.

        The date is snapped to the Sunday of the report week — if report_date is
        already a Sunday it is used as-is, otherwise we go back to the most
        recent Sunday.  Deduplication is by description + date so re-running
        the same week never double-writes.
        """
        if "Income & Exp" not in self.workbook.sheetnames:
            return

        ws = self.workbook["Income & Exp"]

        # ── 1. Snap date to Sunday ─────────────────────────────────────
        weekday = report_date.weekday()   # 0=Mon … 6=Sun
        if weekday == 6:
            sunday = report_date
        else:
            sunday = report_date - timedelta(days=(weekday + 1) % 7)
        sunday = sunday.replace(hour=0, minute=0, second=0, microsecond=0)

        # ── 2. Find last real data row ────────────────────────────────
        last_row = 6
        for r in range(6, ws.max_row + 1):
            b = ws.cell(row=r, column=2).value
            if b is not None and str(b).strip() not in ('', '.'):
                last_row = r

        # ── 3. Dedup: skip if a Contribution Collected row for this
        #    Sunday already exists (description + date match) ──────────
        sunday_date = sunday.date()
        for r in range(6, last_row + 1):
            desc = str(ws.cell(row=r, column=2).value or '').strip().upper()
            cell_date = ws.cell(row=r, column=4).value
            if isinstance(cell_date, datetime):
                cell_date = cell_date.date()
            if desc == 'CONTRIBUTION COLLECTED' and cell_date == sunday_date:
                logger.info("Income & Exp: Contribution Collected row already exists — skipping.")
                return

        # ── 4. Copy style from the last real data row ─────────────────
        def _copy_style(src_row: int, dst_row: int, col: int):
            from openpyxl.styles import Font, Border, Alignment
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=dst_row, column=col)
            if src.has_style:
                dst.font      = src.font.copy()
                dst.border    = src.border.copy()
                dst.alignment = src.alignment.copy()

        # ── 5. Write the new income row ───────────────────────────────
        new_row = last_row + 1

        # Col A: auto-number formula
        s_ref_row = new_row - 6 + 5
        ws.cell(row=new_row, column=1).value = f"=COUNTBLANK($S$5:S{s_ref_row - 1})"

        # Col B: Description
        ws.cell(row=new_row, column=2).value = "CONTRIBUTION COLLECTED"

        # Col C: Income / Expense  (dropdown value — must match list exactly)
        ws.cell(row=new_row, column=3).value = "Income"

        # Col D: Date (the Sunday)
        date_cell = ws.cell(row=new_row, column=4)
        date_cell.value         = sunday
        date_cell.number_format = 'DD-MMM-YY'

        # Col E: Category
        ws.cell(row=new_row, column=5).value = "Contribution"

        # Col F: Amount — positive (income)
        amt_cell = ws.cell(row=new_row, column=6)
        amt_cell.value         = total_amount
        amt_cell.number_format = '#,##0.00'

        # Col G: Running balance
        ws.cell(row=new_row, column=7).value = f"=G{last_row}+F{new_row}"

        # Col H: No receipt number for an aggregated row
        ws.cell(row=new_row, column=8).value = None

        # Copy styles from previous row
        for col in range(1, 9):
            _copy_style(last_row, new_row, col)

        logger.info(
            f"Income & Exp: wrote Contribution Collected "
            f"KES {total_amount:,.2f} for Sunday {sunday.strftime('%d-%b-%Y')}."
        )

    def _update_missing_contro_tab(self, report_date: datetime, members: List[Dict], all_transactions: Optional[List[Dict]] = None):
        """
        Clears and rewrites the 'Missing Contro' sheet with every member
        whose **Contribution** amount this week is 0 or None.

        Uses the members list plus the transaction ledger to compute a
        Contribution-only total, so members who gave Benevolence/Missions
        but no Contribution are still listed as missing.

        Members with a pledge of N/a or blank/non-numeric are excluded.
        """
        # ── 1. Find the Missing Contro sheet ──────────────────────────
        ws_missing = None
        for name in self.workbook.sheetnames:
            if name.strip().lower().startswith('missing contro'):
                ws_missing = self.workbook[name]
                break

        if ws_missing is None:
            logger.warning("No 'Missing Contro' sheet found — skipping.")
            return

        if 'Combined' not in self.workbook.sheetnames:
            logger.warning("No 'Combined' sheet found — skipping Missing Contro update.")
            return

        ws_combined = self.workbook['Combined']

        # ── 2. Build pledge lookup: row_index → pledge value ──────────
        pledge_by_row = {}
        for r in range(2, ws_combined.max_row + 1):
            first = ws_combined.cell(row=r, column=5).value
            if first is None:
                break
            pledge_by_row[r] = ws_combined.cell(row=r, column=7).value

        # ── 3. Build Contribution-only totals per member ─────────────────
        contribution_by_row: Dict[int, float] = {}
        if all_transactions:
            for t in all_transactions:
                category = (t.get('category') or '').strip().lower()
                if category != 'contribution':
                    continue
                member_row = t.get('member_row')
                if member_row is None:
                    continue
                contribution_by_row[member_row] = contribution_by_row.get(member_row, 0.0) + (t.get('amount', 0) or 0)

        # ── 4. Filter members: zero contribution + valid numeric pledge ─
        missing = []
        for m in members:
            pledge = pledge_by_row.get(m.get('row_index'))

            pledge_str = str(pledge or '').strip().lower()
            if pledge_str in ('n/a', 'na', ''):
                continue
            try:
                pledge_val = float(pledge)
            except (TypeError, ValueError):
                continue

            contrib_total = contribution_by_row.get(m.get('row_index'), 0.0)
            if contrib_total == 0:
                missing.append({
                    'first':  str(m.get('first_name', '')).strip(),
                    'last':   str(m.get('last_name',  '')).strip(),
                    'pledge': pledge_val,
                })

        # ── 4. Clear data rows, keep header row 1 ────────────────────
        for r in range(2, ws_missing.max_row + 1):
            for c in range(1, 6):
                ws_missing.cell(row=r, column=c).value = None

        # ── 5. Write new data rows ─────────────────────────────────────
        from openpyxl.styles import Border, Side, Alignment
        _thin        = Side(border_style='thin')
        _full_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        for i, m in enumerate(missing):
            r = i + 2

            seq_cell = ws_missing.cell(row=r, column=1)
            seq_cell.value     = i + 1
            seq_cell.border    = _full_border

            first_cell = ws_missing.cell(row=r, column=2)
            first_cell.value  = m['first']
            first_cell.border = _full_border

            last_cell = ws_missing.cell(row=r, column=3)
            last_cell.value  = m['last']
            last_cell.border = _full_border

            pledge_cell = ws_missing.cell(row=r, column=4)
            pledge_cell.value  = m['pledge']
            pledge_cell.font   = Font(bold=True, size=10)
            pledge_cell.border = _full_border

            given_cell = ws_missing.cell(row=r, column=5)
            given_cell.value  = 0
            given_cell.border = _full_border

        # ── 6. Rename sheet to this report's Sunday ───────────────────
        weekday = report_date.weekday()
        sunday  = report_date if weekday == 6 else report_date - timedelta(days=(weekday + 1) % 7)
        day     = sunday.day
        new_sheet_name = f"Missing Contro {sunday.strftime('%b')} {day}, {sunday.year}"
        try:
            ws_missing.title = new_sheet_name[:31]
        except Exception as e:
            logger.warning(f"Could not rename Missing Contro sheet: {e}")

        logger.info(
            f"Missing Contro: {len(missing)} member(s) with zero contribution "
            f"for {sunday.strftime('%d-%b-%Y')}."
        )

    def _update_missions_tab(self, all_transactions: List[Dict]):
        """
        Appends missions-categorised transactions to the Missions sheet.

        Missions sheet layout (row 1-3 = headers/summary, data from row 4):
          Row 1: Missions Goal | blank | goal_amount | formula
          Row 2: Missions Collected | blank | =SUM(C4:C995) | blank | blank | headers for member list
          Row 3: Date | Name | Amount (Ksh.)
          Row 4+: data rows

        Filters all_transactions for category='Missions', sorts by date,
        deduplicates by receipt_no + amount, and appends new rows.
        """
        if "Missions" not in self.workbook.sheetnames:
            logger.warning("'Missions' sheet not found — skipping.")
            return

        ws = self.workbook["Missions"]

        # ── 1. Find last real data row ───────────────────────────────
        # Scan backward from max_row to find the last non-empty row
        last_row = 3  # fallback to header row
        for r in range(ws.max_row, 3, -1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            if a is not None or b is not None or c is not None:
                last_row = r
                break

        # ── 2. Collect existing dedup keys (receipt_no + amount) ──────
        existing_keys = set()
        for r in range(4, last_row + 1):
            # Missions sheet doesn't have receipt numbers, so use date + name + amount
            date_val = ws.cell(row=r, column=1).value
            name_val = ws.cell(row=r, column=2).value
            amt_val  = ws.cell(row=r, column=3).value
            dedup_key = f"{date_val}|{name_val}|{amt_val}"
            existing_keys.add(dedup_key)

        # ── 3. Filter and sort missions transactions ──────────────────
        missions_txns = [
            t for t in all_transactions
            if (t.get('category') or '').strip().lower() == 'missions'
        ]
        # Sort by date ascending
        missions_txns.sort(key=lambda t: t.get('date') or datetime.min)

        # ── 4. Append each new mission transaction ────────────────────
        new_count = 0
        for t in missions_txns:
            date_val = t.get('date') or datetime.now()
            # Use the matched/converted name, falling back to raw sender_name
            name_val = (t.get('name') or t.get('sender_name') or '').strip()
            # Normalize specific family names for Missions tab display
            name_val = self._normalize_missions_name(name_val)
            amount   = abs(t.get('amount', 0) or 0)

            # Dedup check
            dedup_key = f"{date_val}|{name_val}|{amount}"
            if dedup_key in existing_keys:
                continue

            new_row = last_row + 1

            # Col A: Date
            date_cell = ws.cell(row=new_row, column=1)
            date_cell.value         = date_val
            date_cell.number_format = 'DD-MMM-YY'

            # Col B: Giver Name
            ws.cell(row=new_row, column=2).value = name_val

            # Col C: Amount
            amt_cell = ws.cell(row=new_row, column=3)
            amt_cell.value         = amount
            amt_cell.number_format = '#,##0'

            existing_keys.add(dedup_key)
            last_row  = new_row
            new_count += 1

        logger.info(f"Missions: appended {new_count} new transaction(s).")

    def _normalize_missions_name(self, name: str) -> str:
        """Normalize specific family names for the Missions tab."""
        lower = name.strip().lower()
        if lower == "deji aregbesola":
            return "Aregbesolas"
        if lower in ("james opwondi", "risper opwondi"):
            return "Opwondis"
        return name

    def _create_breakdown_tab(self, report_date: datetime, all_transactions: Optional[List[Dict]] = None, members: Optional[List[Dict]] = None):
        """
        Creates/overwrites the 'Breakdown of Available Funds' tab.

        Layout (matching the church template):
          Row 1:  'Total Balances' (merged A1:E1) | 'Missions' (merged G1:K1)
          Row 2:  Date | Mpesa | Cash | (blank) | Grand Total | Date | Mpesa ZIIDI (formula) | Cash In Hand (running total) | Bank Account | Total Cash (formula)
          Row 3:  (colored boxes)
          Row 5:  'Contribution' header (merged A5:E5) | 'Benevolence' header (merged G5:K5)
          Row 6:  Date | Mpesa Mobile Money | Cash In Hand | Bank Account | Total Cash | Date | Mpesa Mobile Money (ZIIDI) | Cash In Hand | Bank Account | Total Cash
          Row 7:  date row with values
        """
        # ── 1. Find or create the Breakdown sheet ─────────────────────
        sheet_name = f"{report_date.strftime('%d %B, %Y')} Breakdown of Available Funds"
        # Truncate to Excel's 31-char limit if needed
        sheet_name = sheet_name[:31]

        # Remove existing sheet with same name if present
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]

        ws = self.workbook.create_sheet(sheet_name)

        # ── 2. Compute category totals from all_transactions ──────────
        def _cat_total(category_name: str) -> float:
            if not all_transactions:
                return 0.0
            return sum(
                abs(t.get('amount', 0) or 0)
                for t in all_transactions
                if (t.get('category') or '').strip().lower() == category_name.lower()
            )

        contribution_mpesa = _cat_total('Contribution')
        contribution_cash  = _cat_total('Contribution')
        benevolence_mpesa  = _cat_total('Benevolence')
        benevolence_cash   = _cat_total('Benevolence')
        missions_cash      = _cat_total('Missions')

        # For Contribution/Benevolence, try to split Mpesa vs Cash from matched/unmatched
        cont_mpesa = 0.0
        cont_cash  = 0.0
        bene_mpesa = 0.0
        bene_cash  = 0.0
        if all_transactions:
            for t in all_transactions:
                cat = (t.get('category') or '').strip().lower()
                amt = abs(t.get('amount', 0) or 0)
                src = (t.get('source') or '').strip().lower()
                if cat == 'contribution':
                    if src == 'cash':
                        cont_cash += amt
                    else:
                        cont_mpesa += amt
                elif cat == 'benevolence':
                    if src == 'cash':
                        bene_cash += amt
                    else:
                        bene_mpesa += amt

        # If split didn't populate, fall back to totals
        if cont_mpesa == 0 and cont_cash == 0 and contribution_mpesa > 0:
            cont_mpesa = contribution_mpesa
            cont_cash  = contribution_mpesa
        if bene_mpesa == 0 and bene_cash == 0 and benevolence_mpesa > 0:
            bene_mpesa = benevolence_mpesa
            bene_cash  = benevolence_mpesa

        # ── 3. Read previous Breakdown sheet for Cash In Hand carry-forward ──
        prev_cash_in_hand = 0.0
        prev_sheet = None
        for name in self.workbook.sheetnames:
            if 'breakdown' in name.lower() and name != sheet_name:
                prev_sheet = name
                break

        if prev_sheet:
            try:
                ws_prev = self.workbook[prev_sheet]
                # Cash In Hand under Missions block is in column H (col 8), row 7 (first data row)
                # We scan for the Missions block header to locate the right cell
                for r in range(1, 20):
                    val = ws_prev.cell(row=r, column=7).value  # col G
                    if val and 'missions' in str(val).lower():
                        # Cash In Hand is typically col H (8) in the row below header
                        prev_cash_in_hand = ws_prev.cell(row=r+1, column=8).value or 0.0
                        break
            except Exception:
                prev_cash_in_hand = 0.0

        new_cash_in_hand = prev_cash_in_hand + missions_cash

        # ── 4. Write Total Balances section ───────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.cell(row=1, column=1, value="Total Balances")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
        ws.cell(row=1, column=7, value="Missions")
        ws.cell(row=1, column=7).font = Font(bold=True, size=12)
        ws.cell(row=1, column=7).alignment = Alignment(horizontal='center')

        # Row 2: Date | Mpesa | Cash | (blank) | Grand Total | Date | Mpesa ZIIDI | Cash In Hand | Bank Account | Total Cash
        date_str = report_date.strftime("%d-%b-%Y")
        ws.cell(row=2, column=1, value=date_str)
        ws.cell(row=2, column=2, value=cont_mpesa)
        ws.cell(row=2, column=3, value=cont_cash)
        ws.cell(row=2, column=5, value="=SUM(B2:C2)")  # Grand Total formula

        # Missions block row 2
        ws.cell(row=2, column=7, value=date_str)
        # Mpesa ZIIDI: copy formula from previous sheet if available, else use the known formula
        mpesa_ziidi_formula = "='Income & Exp'!C3 + Missions!C2 - SUMIF('Income & Exp'!E:E,\"Missions Transfer\",'Income & Exp'!F:F)"
        ws.cell(row=2, column=8, value=mpesa_ziidi_formula)
        ws.cell(row=2, column=9, value=new_cash_in_hand)
        ws.cell(row=2, column=11, value="=H2+I2")  # Total Cash formula

        # Row 3: colored boxes (green for Mpesa, red for Cash, blue for Total)
        ws.cell(row=3, column=2).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        ws.cell(row=3, column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        ws.cell(row=3, column=5).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.cell(row=3, column=8).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        ws.cell(row=3, column=9).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        ws.cell(row=3, column=11).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # ── 5. Contribution block header ──────────────────────────────
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
        ws.cell(row=5, column=1, value="Contribution")
        ws.cell(row=5, column=1).font = Font(bold=True, size=11)
        ws.cell(row=5, column=1).alignment = Alignment(horizontal='center')

        # Benevolence block header
        ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=11)
        ws.cell(row=5, column=7, value="Benevolence")
        ws.cell(row=5, column=7).font = Font(bold=True, size=11)
        ws.cell(row=5, column=7).alignment = Alignment(horizontal='center')

        # Row 6 sub-headers
        for col, label in [(1, "Date"), (2, "Mpesa Mobile Money"), (3, "Cash In Hand"), (4, "Bank Account"), (5, "Total Cash")]:
            ws.cell(row=6, column=col, value=label).font = Font(bold=True)
        for col, label in [(7, "Date"), (8, "Mpesa Mobile Money (ZIIDI)"), (9, "Cash In Hand"), (10, "Bank Account"), (11, "Total Cash")]:
            ws.cell(row=6, column=col, value=label).font = Font(bold=True)

        # Row 7: data
        ws.cell(row=7, column=1, value=date_str)
        ws.cell(row=7, column=2, value=cont_mpesa)
        ws.cell(row=7, column=3, value=cont_cash)
        ws.cell(row=7, column=4, value=0)
        ws.cell(row=7, column=5, value="=B7+C7")

        ws.cell(row=7, column=7, value=date_str)
        ws.cell(row=7, column=8, value=bene_mpesa)
        ws.cell(row=7, column=9, value=bene_cash)
        ws.cell(row=7, column=10, value=0)
        ws.cell(row=7, column=11, value="=H7+I7")

        # Apply borders to data area
        for r in [2, 3, 7]:
            for c in range(1, 12):
                ws.cell(row=r, column=c).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        # Number formatting
        for r in [2, 7]:
            for c in [2, 3, 5, 8, 9, 11]:
                ws.cell(row=r, column=c).number_format = '#,##0.00'

        logger.info(f"Created Breakdown tab: {sheet_name}")

    def _update_benevolence_tab(self, all_transactions: List[Dict], report_date: datetime):
        """
        Appends benevolence income and expenses to the 'Benevolence ' sheet.

        Sheet layout (row 3 = headers, data from row 4):
          Col A: Date             (format: d"-"mmm"-"yy)
          Col B: Collected        (income — positive, benevolence category only, no transfers)
          Col C: Paid out         (expenses — negative, Benevolence Expense + Benevolence Expense Transaction Charge)
          Col D: Balance          (formula =D(prev)+B(n)  or  =D(prev)+C(n))
          Col E: Special Notes    (transaction details for expense rows)

        Income date is snapped to the last/current Wednesday (benevolence collection day).
        Deduplication uses date + amount + description as composite key.
        """
        # Find the Benevolence sheet (has a trailing space in the name)
        ws = None
        for name in self.workbook.sheetnames:
            if name.strip().lower() == 'benevolence':
                ws = self.workbook[name]
                break
        if ws is None:
            logger.warning("'Benevolence' sheet not found — skipping.")
            return

        DATE_FMT   = 'd"-"mmm"-"yy'
        MONEY_FMT  = '[$Ksh]#,##0.00'

        # ── 1. Find last real data row (scan backward) ────────────────
        last_row = 3  # fallback to header row
        for r in range(ws.max_row, 3, -1):
            if any(ws.cell(row=r, column=c).value is not None for c in range(1, 5)):
                last_row = r
                break

        # ── 2. Collect existing dedup keys ────────────────────────────
        existing_keys = set()
        for r in range(4, last_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            e = ws.cell(row=r, column=5).value
            key = f"{a}|{b}|{c}|{e}"
            existing_keys.add(key)

        # ── 3. Snap report_date to Wednesday ─────────────────────────
        # weekday(): 0=Mon … 2=Wed … 6=Sun
        weekday = report_date.weekday()
        days_since_wed = (weekday - 2) % 7
        wednesday = (report_date - timedelta(days=days_since_wed)).replace(
            hour=0, minute=0, second=0, microsecond=0)

        # ── 4. Aggregate benevolence income (exclude transfers) ───────
        benevolence_income = sum(
            abs(t.get('amount', 0) or 0)
            for t in all_transactions
            if (t.get('category') or '').strip().lower() == 'benevolence'
        )

        # ── 5. Collect expense rows (Benevolence Expense + charge) ───
        BENEV_EXP_CATS = {'benevolence expense', 'benevolence expense transaction charge'}
        expense_txns = [
            t for t in all_transactions
            if (t.get('category') or '').strip().lower() in BENEV_EXP_CATS
        ]
        expense_txns.sort(key=lambda t: t.get('date') or datetime.min)

        def _write_row(new_row, col_b, col_c, note, date_val):
            """Write one row to the sheet."""
            # Col A: Date
            date_cell = ws.cell(row=new_row, column=1)
            date_cell.value         = date_val
            date_cell.number_format = DATE_FMT

            # Col B: Collected (income)
            b_cell = ws.cell(row=new_row, column=2)
            b_cell.value         = col_b
            b_cell.number_format = MONEY_FMT

            # Col C: Paid out (expense, negative)
            c_cell = ws.cell(row=new_row, column=3)
            c_cell.value         = col_c
            c_cell.number_format = MONEY_FMT

            # Col D: Running balance formula
            if col_b is not None:
                ws.cell(row=new_row, column=4).value = f"=D{new_row - 1}+B{new_row}"
            else:
                ws.cell(row=new_row, column=4).value = f"=D{new_row - 1}+C{new_row}"
            ws.cell(row=new_row, column=4).number_format = MONEY_FMT

            # Col E: Special Notes (details for expense rows)
            if note:
                ws.cell(row=new_row, column=5).value = note

        # ── 6. Build all rows to write, then sort by date ────────────
        rows_to_write = []

        if benevolence_income > 0:
            income_key = f"{wednesday}|{benevolence_income}|None|None"
            if income_key not in existing_keys:
                rows_to_write.append({
                    'date':    wednesday,
                    'col_b':   benevolence_income,
                    'col_c':   None,
                    'note':    None,
                    'key':     income_key,
                })

        for t in expense_txns:
            amount   = -abs(t.get('amount', 0) or 0)
            details  = (t.get('details') or t.get('notes') or t.get('sender_name') or '').strip()
            date_val = t.get('date') or wednesday
            exp_key  = f"{date_val}|None|{amount}|{details}"
            if exp_key not in existing_keys:
                rows_to_write.append({
                    'date':    date_val,
                    'col_b':   None,
                    'col_c':   amount,
                    'note':    details,
                    'key':     exp_key,
                })

        # Sort all new rows by date ascending before writing
        rows_to_write.sort(key=lambda r: r['date'] or datetime.min)

        # ── 7. Write sorted rows ──────────────────────────────────────
        for row_data in rows_to_write:
            last_row += 1
            _write_row(last_row, row_data['col_b'], row_data['col_c'],
                       row_data['note'], row_data['date'])
            existing_keys.add(row_data['key'])

        logger.info(f"Benevolence: wrote {len(rows_to_write)} new row(s).")

    def _safe_set_value(self, ws: Worksheet, row: int, col: int, value: any):
        """Safely sets a cell value, skipping MergedCells"""
        from openpyxl.cell.cell import MergedCell
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = value

    def create_report_with_matches(
        self,
        report_date: datetime,
        output_path: str,
        matched_mpesa: List[Dict],
        unmatched_mpesa: List[Dict],
        matched_cash: List[Dict],
        unmatched_cash: List[Dict],
        members: List[Dict]
    ) -> str:
        """
        Populates the workbook using pre-matched data by updating the template sheet in-place.
        Surgically clears dynamic data first, then fills new data and renames the sheet.
        """
        sunday_name = report_date.strftime("%b %d").replace(" 0", " ")
        
        # 1. Get the template sheet
        ws = self._get_template_worksheet()
        if not ws:
            return self._create_from_scratch(report_date, output_path, matched_mpesa, unmatched_mpesa, matched_cash, unmatched_cash, members)

        # 2. Update Date Headers (A4:D6 area)
        if ws['B5'].data_type != 'f':
            self._safe_set_value(ws, 5, 2, report_date)
            ws.cell(row=5, column=2).number_format = 'dd-mmm'
        
        # 3. Clear dynamic data in the Member/Visitor/Transfers area (Rows 10 to 60)
        # We increase the range to ensure no old data (like leftovers from the template) remains.
        for row in range(10, 61):
            # Clear all data columns from F (6) to U (21)
            for col in range(6, 22):
                cell = ws.cell(row=row, column=col)
                
                # 1. Preserve Formulas (like Totals)
                if cell.data_type == 'f':
                    continue
                
                # 2. Preserve Whitelisted Labels (Headers like "Cash Guests", "Benevolence", etc.)
                if col in [10, 13, 16, 18, 20]: # J, M, P, R, T
                    val = cell.value
                    if val and isinstance(val, str):
                        clean_val = val.strip().lower()
                        whitelist = [
                            "cash guests", "cash from guests", "benevolence", "missions transfer", 
                            "none", "original", "check (contro)", 
                            "total missions collected", "benevolence cash", "missions cash", "total",
                            "cross check", "benevolence transfer", "total missions"
                        ]
                        
                        # J10/J11 is usually the "Cash Guests" header
                        if col == 10 and (row == 10 or row == 11) and ("guests" in clean_val or clean_val == "guest"):
                            continue

                        # Preserve anything in the whitelist or starting with an asterisk (explanatory notes)
                        if clean_val.startswith("*") or any(w in clean_val for w in whitelist):
                            continue
                
                # 3. Special Case: Preserve certain static reference areas (S11-S16) if needed
                if col == 19 and 11 <= row <= 16:
                    # Check if it looks like a label
                    if cell.value and isinstance(cell.value, str):
                        continue

                # 4. Preserve Denomination Labels (P35-P46)
                if col == 16 and 35 <= row <= 46:
                    continue

                # Clear everything else
                self._safe_set_value(ws, row, col, None)

        # 4. Fill Member Data (Matched)
        # We rewrite names and ministry to ensure the Weekly list matches the current master ledger
        # Weekly Layout: B=Ministry, C=First Name, D=Last Name, E=Pledge
        for m in members:
            # Calculate the corresponding row in the Weekly sheet (Fixed range starts at 10)
            weekly_row = 10 + (m['row_index'] - 2)
            
            # Apply borders to the extended member area
            for c in range(1, 9):
                ws.cell(row=weekly_row, column=c).border = self.styles['border']
                
            # A (# Number)
            self._safe_set_value(ws, weekly_row, 1, m['row_index'] - 1)
            ws.cell(row=weekly_row, column=1).alignment = self.styles['center']
            self._safe_set_value(ws, weekly_row, 2, m['ministry'])   # B (Ministry)
            self._safe_set_value(ws, weekly_row, 3, m['first_name']) # C (First Name)
            self._safe_set_value(ws, weekly_row, 4, m['last_name'])  # D (Last Name)
            self._safe_set_value(ws, weekly_row, 5, m['pledge'])     # E (Pledge)
            
            # Sum by category (filtering by the member's row in the Master list)
            master_row = m['row_index']
            cont_total = sum(t.get('amount', 0) for t in matched_mpesa if t.get('member_row') == master_row and str(t.get('category', '')).lower().startswith('cont'))
            cont_total += sum(c.get('amount', 0) for c in matched_cash if c.get('member_row') == master_row and str(c.get('category', '')).lower().startswith('cont'))
            
            miss_total = sum(t.get('amount', 0) for t in matched_mpesa if t.get('member_row') == master_row and str(t.get('category', '')).lower().startswith('miss'))
            miss_total += sum(c.get('amount', 0) for c in matched_cash if c.get('member_row') == master_row and str(c.get('category', '')).lower().startswith('miss'))
            
            if cont_total > 0:
                self._safe_set_value(ws, weekly_row, 6, cont_total) # F
            if miss_total > 0:
                self._safe_set_value(ws, weekly_row, 7, miss_total) # G
            
            # Calculate and fill Total (H)
            grand_total = cont_total + miss_total
            if grand_total > 0:
                self._safe_set_value(ws, weekly_row, 8, grand_total) # H

        # 5. Fill Visitors (Column J-K)
        # We consolidate unmatched M-Pesa and Cash, removing duplicates and empty names
        visitor_list = []
        seen_names = set()
        
        # Special handling for "Cash Guest" aggregate total (goes to K11)
        cash_guest_total = 0
        
        for v in unmatched_mpesa + unmatched_cash:
            name = (v.get('name') or "").strip()
            amount = float(v.get('amount', 0))
            if not name or amount <= 0: continue
            
            # Skip benevolence and missions (J & K is only for visitors contribution)
            cat = str(v.get('category') or '').strip().lower()
            if 'benev' in cat or 'miss' in cat:
                continue
            
            # Detect aggregate summary labels from OCR
            lower_name = name.lower()
            is_summary = any(label in lower_name for label in ["cash guest", "guest", "missions cash", "benevolence cash"])
            if is_summary:
                if "guest" in lower_name:
                    cash_guest_total += amount
                continue
            
            # Use lowercase for duplicate detection but keep original for display
            key = name.lower()
            if key not in seen_names:
                seen_names.add(key)
                visitor_list.append({'name': name, 'amount': amount, 'category': v.get('category', '')})

        # Plug Cash Guest total into K11 (next to the Cash Guest label)
        if cash_guest_total > 0:
            self._safe_set_value(ws, 11, 11, cash_guest_total)

        j_row = 11
        for v in visitor_list:
            if j_row > 60: break
            
            # Skip row 29 ONLY if it's a formula/protected row (to avoid gaps in a plain list)
            cell_j = ws.cell(row=j_row, column=10)
            if j_row == 29 and cell_j.data_type == 'f':
                j_row += 1
            
            # Don't overwrite the header if we are at row 11 or 12
            curr_val = str(ws.cell(row=j_row, column=10).value or "").lower()
            if "guests" in curr_val and j_row <= 12:
                j_row += 1
                
            self._safe_set_value(ws, j_row, 10, v['name'])
            self._safe_set_value(ws, j_row, 11, v['amount'])
            j_row += 1

        # 6. Fill Mobile Money Transfers (Column P-Q and Overflow into T-U)
        # Only pure Contribution entries — no Missions, no Benevolence
        def _is_contribution_only(entry):
            cat = str(entry.get('category', '')).strip().lower()
            return cat not in ('missions', 'benevolence') and not cat.startswith('miss') and not cat.startswith('benev')

        mobile_entries_map = {}
        for t in matched_mpesa:
            if _is_contribution_only(t):
                name = t['name']
                mobile_entries_map[name] = mobile_entries_map.get(name, 0.0) + float(t.get('amount', 0))

        # Add "GUEST MPESA" total from visitors who gave Contribution only
        guest_mpesa_total = sum(v['amount'] for v in visitor_list if _is_contribution_only(v))
        if guest_mpesa_total > 0:
            mobile_entries_map["GUEST MPESA"] = mobile_entries_map.get("GUEST MPESA", 0.0) + guest_mpesa_total

        mobile_entries = [{'name': name, 'amount': amt} for name, amt in mobile_entries_map.items()]

        curr_row = 11
        curr_col_name = 16 # P
        curr_col_amt = 17 # Q
        
        for t in mobile_entries:
            # If we reached the end of P-Q area (Row 28), overflow to T-U
            if curr_row >= 29:
                if curr_col_name == 16: # Switch from P-Q to T-U
                    curr_col_name = 20 # T
                    curr_col_amt = 21 # U
                    curr_row = 11
                else:
                    break # Out of space in both sections
            
            self._safe_set_value(ws, curr_row, curr_col_name, t['name'])
            self._safe_set_value(ws, curr_row, curr_col_amt, t['amount'])
            curr_row += 1

        # 7. Fill "Others" Summary section (Column R-S, rows 11-16)
        # S11: Benevolence Transfer, S13: Total Missions, S14: Benevolence Cash, S15: Missions Cash, S16: Missions Transfer
        
        # 1. Calculate Transfers from M-Pesa
        b_trans = sum(t['amount'] for t in matched_mpesa + unmatched_mpesa if t.get('category') == "Benevolence")
        m_trans = sum(t['amount'] for t in matched_mpesa + unmatched_mpesa if t.get('category') == "Missions")
        
        # 2. Calculate Cash from Images
        # We prioritize explicit "Total" labels (like "Benevolence Cash Total") if they exist on the image
        b_total_label = next((c['amount'] for c in matched_cash + unmatched_cash if "benevolence cash" in c['name'].lower()), None)
        m_total_label = next((c['amount'] for c in matched_cash + unmatched_cash if "missions cash" in c['name'].lower()), None)
        
        if b_total_label is not None:
            b_cash = b_total_label
        else:
            b_cash = sum(c['amount'] for c in matched_cash + unmatched_cash if c.get('category') == "Benevolence")

        if m_total_label is not None:
            m_cash = m_total_label
        else:
            m_cash = sum(c['amount'] for c in matched_cash + unmatched_cash if c.get('category') == "Missions")

        # 3. Write to Excel
        self._safe_set_value(ws, 11, 19, b_trans) # S11
        self._safe_set_value(ws, 14, 19, b_cash)  # S14
        self._safe_set_value(ws, 15, 19, m_cash)  # S15
        self._safe_set_value(ws, 16, 19, m_trans) # S16
        
        # Set formula for Total Missions
        ws.cell(row=13, column=19, value="=S15+S16")

        # 8. Rename the sheet to the new date
        ws.title = sunday_name
        
        # 9. Fill Cash Breakdown (Rows 35-43 approx)
        if hasattr(self, 'cash_breakdown') and self.cash_breakdown:
            self._fill_cash_breakdown(ws, self.cash_breakdown)

        # 8. Rename the sheet to the new date
        ws.title = sunday_name
        
        # 9. Fill Cash Breakdown (Rows 35-43 approx)
        if hasattr(self, 'cash_breakdown') and self.cash_breakdown:
            self._fill_cash_breakdown(ws, self.cash_breakdown)

        # NOTE: _update_combined_sheet and _update_logs are now handled exclusively by finalize_report
        # to avoid double-processing.
        
        return output_path

    def _fill_cash_breakdown(self, ws: Worksheet, breakdown: Dict[float, float]):
        """
        Fills the cash breakdown section (denominations)
        """
        # Search Column P (16) for denominations (Rows 34-50)
        for row in range(34, 50):
            label_cell = ws.cell(row=row, column=16)
            denom_val = label_cell.value
            
            if denom_val and isinstance(denom_val, (int, float)):
                # If we have a total for this denomination, put it in Column Q (17)
                if denom_val in breakdown:
                    self._safe_set_value(ws, row, 17, breakdown[denom_val])
                elif float(denom_val) in breakdown:
                    self._safe_set_value(ws, row, 17, breakdown[float(denom_val)])

    def _get_template_worksheet(self) -> Optional[Worksheet]:
        """Finds the most recent Sunday sheet to use as a template"""
        # Look for sheet names like "May 3", "Apr 26", etc.
        # For simplicity, pick the first one that isn't Summary/Combined/Missions
        excluded = ['Summary', 'Combined', 'Missions', 'Report By Bible Talk', 'Attendance']
        for name in self.workbook.sheetnames:
            if name not in excluded and any(month in name for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
                return self.workbook[name]
        return None

    def _create_from_scratch(self, report_date, output_path, matched_mpesa, unmatched_mpesa, matched_cash, unmatched_cash, members):
        """Original implementation as fallback"""
        # (Keeping the previous implementation here or just failing)
        ws = self.workbook.create_sheet(report_date.strftime("%b %d"), 0)
        # ... (rest of old logic)
        self.workbook.save(output_path)
        return output_path
    def _update_logs(self, report_date: datetime, mpesa: List[Dict], cash: List[Dict]):
        """Updates Missions, Benevolence and other logs"""
        self._update_missions_log(report_date, mpesa, cash)
        self._update_benevolence_log(report_date, mpesa, cash)

    def _update_missions_log(self, report_date: datetime, mpesa: List[Dict], cash: List[Dict]):
        """Appends missions entries to the Missions log"""
        if "Missions" not in self.workbook.sheetnames:
            return
        
        ws = self.workbook["Missions"]
        # Find first empty row (looking at Column A, starting from row 4)
        next_row = 4
        while ws.cell(row=next_row, column=1).value:
            next_row += 1
            
        # Append Missions entries from M-Pesa
        for t in mpesa:
            if t.get('category') == "Missions":
                ws.cell(row=next_row, column=1, value=report_date)
                ws.cell(row=next_row, column=1).number_format = 'yyyy-mm-dd'
                ws.cell(row=next_row, column=2, value=t['name'])
                ws.cell(row=next_row, column=3, value=t['amount'])
                next_row += 1
                
        # Append Missions entries from Cash
        for c in cash:
            if c.get('category') == "Missions":
                ws.cell(row=next_row, column=1, value=report_date)
                ws.cell(row=next_row, column=1).number_format = 'yyyy-mm-dd'
                ws.cell(row=next_row, column=2, value=c['name'])
                ws.cell(row=next_row, column=3, value=c['amount'])
                next_row += 1

    def _update_benevolence_log(self, report_date: datetime, mpesa: List[Dict], cash: List[Dict]):
        """Appends benevolence entries to the Benevolence log"""
        if "Benevolence" not in self.workbook.sheetnames:
            return
            
        ws = self.workbook["Benevolence"]
        # Find first empty row (looking at Column A, starting from row 4)
        next_row = 4
        while ws.cell(row=next_row, column=1).value:
            next_row += 1
            
        # Calculate total benevolence for the day
        total_benev = sum(t['amount'] for t in mpesa if t.get('category') == "Benevolence")
        total_benev += sum(c['amount'] for c in cash if c.get('category') == "Benevolence")
        
        if total_benev > 0:
            ws.cell(row=next_row, column=1, value=report_date)
            ws.cell(row=next_row, column=1).number_format = 'yyyy-mm-dd'
            ws.cell(row=next_row, column=2, value=total_benev)
            # Column C is for Paid Out (empty here)
            # Column D is balance (formula)
            if next_row > 4:
                ws.cell(row=next_row, column=4, value=f"=D{next_row-1}+B{next_row}+C{next_row}")
            ws.cell(row=next_row, column=5, value="Sunday Collection")

    def _initialize_empty_workbook(self):
        """Creates a basic structure if no template is used"""
        ws = self.workbook.active
        ws.title = "Summary"
        self.workbook.create_sheet("Combined")
        self.workbook.create_sheet("Missions")
        self.workbook.create_sheet("Benevolence")

    def _combine_entries(self, mpesa_transactions: List[Dict], handwritten_entries: List[Dict]) -> List[Dict]:
        # Legacy method if needed, but we prefer template-based matching
        return []
