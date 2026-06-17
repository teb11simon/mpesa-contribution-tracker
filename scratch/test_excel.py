import openpyxl
from datetime import datetime
import sys
import os

# Add src/ to python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../src')))

from excel_generator import ExcelGenerator

def main():
    template_path = "C:/Users/User/Downloads/Web App May 10, 2026 Nairobi Contribution Report.xlsx"
    output_path = "output/Test_Generated_Report.xlsx"
    
    print(f"Loading template from: {template_path}")
    generator = ExcelGenerator()
    generator.load_template(template_path)
    
    # We will simulate calling _update_combined_sheet.
    # We need a list of members. We can get them using _get_members_from_combined.
    members = generator._get_members_from_combined()
    print(f"Extracted {len(members)} members.")
    
    # Mock data
    mpesa_tx = [
        {'member_row': 2, 'amount': 250, 'category': 'Contribution', 'sender_name': 'Joshua Abungu', 'name': 'Joshua Abungu'},
        {'member_row': 5, 'amount': 7000, 'category': 'Contribution', 'sender_name': 'Fred Arogo', 'name': 'Fred Arogo'},
        {'member_row': 5, 'amount': 800, 'category': 'Missions', 'sender_name': 'Fred Arogo', 'name': 'Fred Arogo'}
    ]
    
    cash_ex = [
        {'member_row': 3, 'amount': 5100, 'category': 'Contribution', 'name': 'Deji Aregbesola'}
    ]
    
    # Mock unmatched transactions (visitors)
    unmatched_mpesa = [
        {'name': 'Visitor One', 'amount': 1000, 'category': 'Contribution'},
        {'name': 'Visitor Two', 'amount': 2000, 'category': 'Missions'},
        {'name': 'Visitor Three', 'amount': 3000, 'category': 'Benevolence'}
    ]
    
    unmatched_cash = [
        {'name': 'Visitor Four', 'amount': 500, 'category': 'Contribution'},
        {'name': 'Visitor Five (Missions)', 'amount': 1500, 'category': 'Missions'}
    ]
    
    # Let's run create_report_with_matches (which creates/updates the weekly sheet)
    report_date = datetime(2026, 5, 24)
    sheet_name = "May 24"
    
    print(f"Running create_report_with_matches for date: {sheet_name}")
    generator.create_report_with_matches(
        report_date=report_date,
        output_path=output_path,
        matched_mpesa=mpesa_tx,
        unmatched_mpesa=unmatched_mpesa,
        matched_cash=cash_ex,
        unmatched_cash=unmatched_cash,
        members=members
    )
    
    print(f"Calling _update_combined_sheet for date: {sheet_name}")
    generator._update_combined_sheet(sheet_name, mpesa_tx, cash_ex, members)
    
    # Save the output workbook
    print(f"Saving generated workbook to: {output_path}")
    generator.workbook.save(output_path)
    print("Workbook saved successfully!")
    
    # Now let's reload it and verify the cells in columns J and K of the weekly sheet
    wb = openpyxl.load_workbook(output_path)
    ws_week = wb[sheet_name]
    
    print("\n--- Verifying Weekly Visitors (Columns J & K) ---")
    print(f"Row 9 Column J (Header): {ws_week.cell(row=9, column=10).value}")
    
    # Read row 11 to 20 on weekly sheet Columns J and K
    for r in range(11, 21):
        v_name = ws_week.cell(row=r, column=10).value
        v_amt = ws_week.cell(row=r, column=11).value
        if v_name or v_amt:
            print(f"Row {r}: Name: {v_name} | Amount: {v_amt}")
            
    print("\n--- Verifying Combined Sheet Row 2 Joshua Abungu ---")
    ws_comb = wb['Combined']
    print(f"Joshua Abungu Formula: {ws_comb.cell(row=2, column=9).value}")

if __name__ == "__main__":
    main()
